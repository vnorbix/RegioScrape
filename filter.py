from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

@dataclass
class ArticleFilter:
    categories: set[str]
    article_numbers: set[str]

def load_filter_data(file: Path) -> ArticleFilter:
    wb = load_workbook(filename=file)
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]

    article_number_idx = headers.index("Cikkszám")
    article_category_idx = headers.index("Főkategória")

    article_filter = ArticleFilter(categories=set(), article_numbers=set())
    for row in sheet.iter_rows(min_row=2, values_only=True):
        article_filter.categories.add(row[article_category_idx])
        article_filter.article_numbers.add(f"{int(row[article_number_idx]):05d}")
    return article_filter