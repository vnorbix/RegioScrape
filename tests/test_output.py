import os
import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook

from output import save_products
from product import Product


@pytest.fixture
def product():
    return Product(article_number='09366',
                   description='&lt;p&gt;&lt;strong&gt;Bburago 1 /43 F1 versenyautó - Red Bull RB19 #1(Max Verstappen)&lt;/strong&gt;&lt;/p&gt; &lt;p&gt;\xa0&lt;/p&gt; &lt;p&gt;Fedezd fel a Bburago 1/43 méretarányú F1 versenyautóját, amely a Red Bull Racing RB19-et ábrázolja, Max Verstappen #1-es számú autóját. Ez a részletesen kidolgozott modell tökéletes választás minden Forma-1 rajongó és gyűjtő számára. A Bburago híres a minőségi és élethű modelleiről, és ez az RB19 sem kivétel. A precíz festés és az aprólékos részletek hűen tükrözik az eredeti versenyautót, amely 2024-ben dominálta a pályákat. Szerezd be most ezt a lenyűgöző modellt, és éld át újra a versenyek izgalmát otthonodban!&lt;/p&gt;',
                   images=['https://www.regiojatek.hu/data/regio_images/normal2/09366_0.jpg',
                    'https://www.regiojatek.hu/data/regio_images/normal2/09366_1.jpg',
                    'https://www.regiojatek.hu/data/regio_images/normal2/09366_2.jpg',
                    'https://www.regiojatek.hu/data/regio_images/normal2/09366_3.jpg',
                    'https://www.regiojatek.hu/data/regio_images/normal2/09366_4.jpg',
                    'https://www.regiojatek.hu/data/regio_images/normal2/09366_5.jpg',
                    'https://www.regiojatek.hu/data/regio_images/normal2/09366_6.jpg',
                    'https://www.regiojatek.hu/data/regio_images/normal2/09366_7.jpg'],
                   age_min=12,
                   age_max=99,
                   gender="unisex")

def test_save_data(product):
    output_path = Path(__file__).parent / "Bburago_termekek_output.xlsx"
    try:
        shutil.copy(Path(__file__).parent / "Bburago_termekek.xlsx", output_path)
        save_products(output_path, [product])

        wb = load_workbook(filename=output_path)
        sheet = wb.active
        assert sheet.cell(row=105, column=11).value == product.age_min
        assert sheet.cell(row=105, column=12).value == product.age_max
        assert sheet.cell(row=105, column=13).value == product.gender
        assert sheet.cell(row=105, column=14).value == product.description
        assert sheet.cell(row=105, column=15).value == ", ".join(product.images)
    finally:
        os.remove(output_path)
        pass



