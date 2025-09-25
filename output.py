import logging
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from product import Product

logger = logging.getLogger(__name__)

def save_products(file: Path, products: Iterable[Product]) -> None:
    wb = load_workbook(filename=file)
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]

    headers_to_save = ["Életkortól", "Életkorig", "Nem", "Leírás", "Képek/Videók"]

    article_number_idx = headers.index("Cikkszám")
    try:
        column_indexes = [headers.index(header) for header in headers_to_save]
    except ValueError:
        next_column_index = _find_next_empty_column(headers)
        column_indexes = [i for i in (range(next_column_index, next_column_index + len(headers_to_save)))]

    for header, column_index in zip(headers_to_save, column_indexes):
        sheet.cell(row=1, column=column_index).value = header

    products_by_article_number = {product.article_number: product for product in products}
    for row in sheet.iter_rows(min_row=2):
        article_number = f"{int(row[article_number_idx].value):05d}"
        if article_number in products_by_article_number:
            product = products_by_article_number[article_number]
            for column_index, value in zip(column_indexes, [product.age_min, product.age_max, product.gender, product.description, ", ".join(product.images)]):
                row[column_index - 1].value = value
        else:
            logger.warning("Az alábbi terméket nem sikerült letölteni a bemeneti listából: %s", article_number)
    wb.save(file)


def _find_next_empty_column(headers) -> int:
    next_empty_column_index = len(headers)
    for i, header in enumerate(reversed(headers)):
        if header is not None:
            next_empty_column_index = len(headers) - i + 1
            break
    return next_empty_column_index
